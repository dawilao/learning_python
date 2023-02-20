import sys
import openpyxl
import xlrd
from unidecode import unidecode

file_path = input("Informe o local e o nome do arquivo: ")
try:
    workbook = openpyxl.load_workbook(file_path)
except openpyxl.utils.exceptions.InvalidFileException:
    try:
        xls_book = xlrd.open_workbook(file_path)
    except FileNotFoundError:
        print("O arquivo não foi encontrado.")
        sys.exit()
    except xlrd.biffh.XLRDError:
        print("Ocorreu um erro ao abrir o arquivo. Verifique se o arquivo está corrompido ou se você tem permissão para abri-lo.")
        sys.exit()
    workbook = openpyxl.Workbook()
    for sheet_index in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        sheet = workbook.create_sheet(xls_sheet.name)
        for row_index in range(xls_sheet.nrows):
            row = []
            for cell in xls_sheet.row(row_index):
                row.append(cell.value)
            sheet.append(row)
    workbook.save(file_path.rsplit(".", 1)[0] + ".xlsx")
    workbook = openpyxl.load_workbook(file_path.rsplit(".", 1)[0] + ".xlsx")

sheets = workbook.sheetnames

for sheet_name in sheets:
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None or isinstance(cell.value, int) or isinstance(cell.value, float):
                continue
            cell.value = unidecode(cell.value).replace("ç", "c")

workbook.save(file_path.rsplit(".", 1)[0] + "_sem_acento.xlsx")
print("Processo concluído.")